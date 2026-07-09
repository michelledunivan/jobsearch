"""
Daily VP Analytics & Insights Job Search
- Searches jobs via Adzuna API (free, 250 requests/day)
- Emails a formatted Excel report every morning
- No AI, no resume, no cost
"""

import os
import time
import smtplib
import requests
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime, timezone, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config — edit these to match your profile ─────────────────────────────────

# Fit keywords scored against each listing (not technical skills — these
# signal senior analytics-executive scope).
SKILLS = ["Analytics", "Insights", "Strategy", "Leadership", "Executive",
          "Stakeholder", "Data Governance"]

SEARCH_QUERIES = [
    "VP Analytics",
    "Vice President Analytics",
    "VP Insights",
    "Vice President Insights",
    "Head of Analytics",
    "Head of Insights",
    "VP Data",
    "General Manager",
]

# Total-comp floor. Listings with a posted max below this get flagged.
SALARY_FLOOR = 240_000

# Keep a job only if it reads remote OR sits in these locations.
LOCATION_KEYWORDS = ["arizona", "phoenix"]

# Title noise (mostly from the General Manager query) and out-of-scope
# employers (government) to drop outright.
EXCLUDE_TITLE_WORDS = [
    "restaurant", "hotel", "retail", "store", "dealership", "automotive",
    "franchise", "gym", "salon", "spa", "warehouse", "car wash",
    "apartment", "property",
]
EXCLUDE_COMPANY_WORDS = [
    "city of", "county of", "state of", "department of", "federal",
]

# ── Constants ──────────────────────────────────────────────────────────────────

OUTPUT_DIR       = Path("output")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
ADZUNA_BASE      = "https://api.adzuna.com/v1/api/jobs/us/search"
HOURS_BACK       = 24
RESULTS_PER_PAGE = 20

# ── Job Fetching ───────────────────────────────────────────────────────────────

def fetch_jobs(app_id: str, app_key: str) -> list[dict]:
    all_jobs, seen_ids = [], set()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=HOURS_BACK)

    for query in SEARCH_QUERIES:
        print(f"  Searching: '{query}' ...")
        for page in range(1, 4):
            params = {
                "app_id": app_id, "app_key": app_key, "what": query,
                "results_per_page": RESULTS_PER_PAGE,
                "sort_by": "date", "max_days_old": 1,
            }
            try:
                resp = requests.get(f"{ADZUNA_BASE}/{page}",
                                    params=params, timeout=30)
                resp.raise_for_status()
                jobs = resp.json().get("results", [])
                if not jobs:
                    break
                added = 0
                for job in jobs:
                    jid = job.get("id")
                    if not jid or jid in seen_ids:
                        continue
                    try:
                        dt = datetime.fromisoformat(
                            job.get("created", "").replace("Z", "+00:00"))
                        if dt < cutoff:
                            continue
                    except Exception:
                        pass
                    seen_ids.add(jid)
                    all_jobs.append(job)
                    added += 1
                print(f"    Page {page}: {added} new jobs (total: {len(all_jobs)})")
                if len(jobs) < RESULTS_PER_PAGE:
                    break
                time.sleep(1)
            except Exception as e:
                print(f"    Warning: page {page} failed - {e}")
                time.sleep(3)
                break

    print(f"  Total unique jobs fetched: {len(all_jobs)}")
    return all_jobs


# ── Data Processing ────────────────────────────────────────────────────────────

def score_job(job: dict) -> int:
    text = ((job.get("title") or "") + " " +
            (job.get("description") or "")).lower()
    return sum(1 for s in SKILLS if s.lower() in text)


def keep_job(title: str, company: str, area_text: str, is_remote: bool) -> bool:
    t = title.lower()
    if any(w in t for w in EXCLUDE_TITLE_WORDS):
        return False
    c = company.lower()
    if any(w in c for w in EXCLUDE_COMPANY_WORDS):
        return False
    return is_remote or any(w in area_text for w in LOCATION_KEYWORDS)


def parse_jobs(raw: list[dict]) -> list[dict]:
    parsed = []
    for job in raw:
        try:
            dt     = datetime.fromisoformat(
                job.get("created", "").replace("Z", "+00:00"))
            posted = dt.strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            posted = "Unknown"

        area     = (job.get("location") or {}).get("area", [])
        location = ", ".join(filter(None, area[-2:])) if area else "Unknown"

        desc      = (job.get("description") or "")[:3000]
        title     = job.get("title") or ""
        company   = (job.get("company") or {}).get("display_name") or ""
        is_remote = "remote" in (title + desc).lower()

        if not keep_job(title, company, " ".join(area).lower(), is_remote):
            continue

        sal_min = job.get("salary_min")
        sal_max = job.get("salary_max")
        salary  = (f"${sal_min:,.0f} - ${sal_max:,.0f}" if sal_min and sal_max
                   else f"${sal_min:,.0f}+" if sal_min else "Not listed")
        if sal_max:
            salary += ("  (meets comp bar)" if sal_max >= SALARY_FLOOR
                       else "  (below comp bar)")

        matched = [s for s in SKILLS if s.lower() in (desc + title).lower()]

        parsed.append({
            "Title":              title,
            "Company":            company,
            "Location":           location,
            "Remote":             "Yes" if is_remote else "No",
            "Salary":             salary,
            "Posted":             posted,
            "Keyword Match":      ", ".join(matched),
            "Match Score":        score_job(job),
            "Key Qualifications": desc[:500],
            "Apply Link":         job.get("redirect_url") or "",
        })

    parsed.sort(key=lambda x: x["Match Score"], reverse=True)
    return parsed


# ── Excel Builder ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F4E79")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0")
HIGH_FILL   = PatternFill("solid", start_color="E2EFDA")

COLUMNS = [
    ("Title",                          35),
    ("Company",                        25),
    ("Location",                       20),
    ("Remote",                          9),
    ("Salary",                         28),
    ("Posted",                         20),
    ("Keyword Match",                  30),
    (f"Match Score (/{len(SKILLS)})",  14),
    ("Key Qualifications",             45),
    ("Apply Link",                     15),
]

SCORE_COL = f"Match Score (/{len(SKILLS)})"


def build_excel(jobs: list[dict], filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title   = "Jobs"
    run_date   = datetime.now().strftime("%B %d, %Y")
    last_col   = get_column_letter(len(COLUMNS))
    thin = Border(
        left=Side(style="thin",   color="BFBFBF"),
        right=Side(style="thin",  color="BFBFBF"),
        top=Side(style="thin",    color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    # Title row
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"VP Analytics & Insights Job Search  —  {run_date}  |  Past 24 Hours"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle row
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (f"Fit keywords: {', '.join(SKILLS)}   |   "
                f"Comp bar: ${SALARY_FLOOR:,}   |   "
                f"Total roles found: {len(jobs)}")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    ws.append([])  # spacer

    # Header row
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for col_idx, (_, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 22

    # Data rows
    for row_num, job in enumerate(jobs, start=5):
        ws.append([job.get(h if h != SCORE_COL else "Match Score", "")
                   for h in headers])

        is_high = job.get("Match Score", 0) >= (len(SKILLS) - 2)
        fill    = HIGH_FILL if is_high else (
                  ALT_FILL  if row_num % 2 == 0 else PatternFill())

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin
            cell.font   = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(col_idx in (1, 9)))
            if fill:
                cell.fill = fill

        # Apply link hyperlink
        apply_col = headers.index("Apply Link") + 1
        url = job.get("Apply Link", "")
        if url:
            c = ws.cell(row=row_num, column=apply_col)
            c.hyperlink = url
            c.font  = Font(name="Arial", size=10, color="0563C1", underline="single")
            c.value = "Apply ->"

        # Center align score and remote
        ws.cell(row=row_num, column=headers.index(SCORE_COL)+1).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_num, column=headers.index("Remote")+1).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 32

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col}4"
    wb.save(filepath)
    print(f"  Excel saved: {filepath.name}")


# ── Email ──────────────────────────────────────────────────────────────────────

def send_email(excel_path: Path, job_count: int,
               gmail_address: str, app_password: str) -> None:
    today   = datetime.now().strftime("%B %d, %Y")
    subject = f"VP Analytics & Insights Jobs — {job_count} Roles Found — {today}"
    body    = f"""Good morning!

Here are the VP / Head of Analytics & Insights roles (plus GM roles)
posted in the last 24 hours, filtered to remote or Phoenix-area.

  Total roles found:       {job_count}
  Fit keywords:            {', '.join(SKILLS)}
  Comp bar:                ${SALARY_FLOOR:,} (flagged in the Salary column)

Open the attached Excel to browse all roles.
Green rows = strongest keyword matches.
Click "Apply ->" in the last column to go directly to the job.

Good luck today!
"""
    msg = MIMEMultipart()
    msg["From"]    = gmail_address
    msg["To"]      = gmail_address
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f"attachment; filename={excel_path.name}")
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(gmail_address, app_password)
        server.sendmail(gmail_address, gmail_address, msg.as_string())
    print(f"  Email sent to {gmail_address}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    adzuna_app_id  = os.environ["ADZUNA_APP_ID"]
    adzuna_app_key = os.environ["ADZUNA_APP_KEY"]
    gmail_address  = os.environ["GMAIL_ADDRESS"]
    app_password   = os.environ["GMAIL_APP_PASSWORD"]

    today     = datetime.now().strftime("%Y-%m-%d")
    out_excel = OUTPUT_DIR / f"VP_Analytics_Jobs_{today}.xlsx"

    print("=== VP Analytics & Insights Job Search ===")
    print(f"Date: {today}\n")

    print("[1/3] Fetching jobs from Adzuna...")
    raw_jobs = fetch_jobs(adzuna_app_id, adzuna_app_key)

    print("\n[2/3] Parsing and scoring jobs...")
    jobs = parse_jobs(raw_jobs)
    print(f"  Jobs parsed: {len(jobs)}")

    print("\n[3/3] Building Excel and sending email...")
    build_excel(jobs, out_excel)
    send_email(out_excel, len(jobs), gmail_address, app_password)

    print(f"\nDone! {len(jobs)} jobs emailed to {gmail_address}")


if __name__ == "__main__":
    main()
